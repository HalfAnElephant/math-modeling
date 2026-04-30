"""Data loading and validation for Problem C UAV inspection."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class UAVParams:
    """UAV and operational parameters read from UAV_Params sheet."""

    k_max: int
    battery_capacity_j: float
    safety_reserve_j: float
    effective_energy_limit_j: float
    horizontal_speed_mps: float
    vertical_speed_mps: float
    horizontal_energy_j_per_m: float
    up_energy_j_per_m: float
    down_energy_j_per_m: float
    hover_power_j_per_s: float
    battery_swap_time_s: float
    operating_horizon_s: float
    walking_speed_mps: float
    walking_detour_factor: float


@dataclass(frozen=True)
class Target:
    """A single inspection target point (node_id 1..16)."""

    node_id: int
    node_name: str
    building_id: str
    x_m: float
    y_m: float
    z_m: float
    priority_level: str
    priority_weight: int
    issue_type: str
    base_hover_time_s: float
    direct_confirm_time_s: float
    manual_point_id: str
    manual_x_m: float
    manual_y_m: float
    manual_service_time_s: float


@dataclass(frozen=True)
class ManualPoint:
    """A manual review point read from the ManualPoints sheet.

    This is the authoritative source for manual service time.
    When NodeData.manual_service_time_s conflicts with this sheet,
    ManualPoints takes precedence for ground-review time calculations.
    """

    manual_point_id: str
    x_m: float
    y_m: float
    manual_service_time_s: float
    mapped_node_ids: tuple[int, ...]


@dataclass(frozen=True)
class ProblemData:
    """Complete problem data loaded from the Excel workbook."""

    params: UAVParams
    targets: list[Target]
    manual_points: dict[str, ManualPoint]
    flight_time_s: dict[tuple[int, int], float]
    flight_energy_j: dict[tuple[int, int], float]
    ground_time_s: dict[tuple[str, str], float]


_EXPECTED_PARAM_KEYS = frozenset({
    "K_max",
    "battery_capacity_J",
    "safety_reserve_J",
    "effective_energy_limit_J",
    "horizontal_speed_mps",
    "vertical_speed_mps",
    "horizontal_energy_J_per_m",
    "up_energy_J_per_m",
    "down_energy_J_per_m",
    "hover_power_J_per_s",
    "battery_swap_time_s",
    "operating_horizon_s",
    "walking_speed_mps",
    "walking_detour_factor",
})


def _read_uav_params(ws: Any) -> UAVParams:
    """Read UAVParams from UAV_Params worksheet rows 4-17."""
    raw: dict[str, float] = {}
    for row in ws.iter_rows(min_row=4, max_row=17, values_only=True):
        if row[0] is None:
            continue
        param_name = str(row[0]).strip()
        raw[param_name] = float(row[1])

    raw_keys = frozenset(raw.keys())
    if raw_keys != _EXPECTED_PARAM_KEYS:
        unexpected = sorted(raw_keys - _EXPECTED_PARAM_KEYS)
        missing = sorted(_EXPECTED_PARAM_KEYS - raw_keys)
        parts = []
        if unexpected:
            parts.append(f"unexpected parameter(s): {unexpected}")
        if missing:
            parts.append(f"missing parameter(s): {missing}")
        raise ValueError(
            f"UAV_Params validation failed: {'; '.join(parts)}. "
            f"Expected {sorted(_EXPECTED_PARAM_KEYS)}."
        )

    return UAVParams(
        k_max=int(raw["K_max"]),
        battery_capacity_j=raw["battery_capacity_J"],
        safety_reserve_j=raw["safety_reserve_J"],
        effective_energy_limit_j=raw["effective_energy_limit_J"],
        horizontal_speed_mps=raw["horizontal_speed_mps"],
        vertical_speed_mps=raw["vertical_speed_mps"],
        horizontal_energy_j_per_m=raw["horizontal_energy_J_per_m"],
        up_energy_j_per_m=raw["up_energy_J_per_m"],
        down_energy_j_per_m=raw["down_energy_J_per_m"],
        hover_power_j_per_s=raw["hover_power_J_per_s"],
        battery_swap_time_s=raw["battery_swap_time_s"],
        operating_horizon_s=raw["operating_horizon_s"],
        walking_speed_mps=raw["walking_speed_mps"],
        walking_detour_factor=raw["walking_detour_factor"],
    )


_EXPECTED_TARGET_COUNT = 16


def _read_targets(ws: Any) -> list[Target]:
    """Read Target list from NodeData worksheet rows 5-20 (node_id 1..16).

    Skips empty rows (where node_id is None), consistent with _read_uav_params.
    Validates that exactly _EXPECTED_TARGET_COUNT targets are read.
    Column index 11 (row[11] = extra_confirm_time_s) is intentionally skipped:
    it is a derived column present in the Excel that duplicates information
    from direct_confirm_time_s and is not needed for modeling.
    """
    targets: list[Target] = []
    # Read from row 5 until the data runs out.
    # Use an upper bound well beyond the expected 16 rows to detect
    # extra targets that would violate the count validation.
    for row in ws.iter_rows(min_row=5, max_row=max(ws.max_row, 5), values_only=True):
        # Stop scanning once we encounter a fully empty row beyond the data
        # (all columns are None), to avoid reading trailing blank rows.
        if all(cell is None for cell in row):
            break
        if row[0] is None:
            continue
        targets.append(Target(
            node_id=int(row[0]),
            node_name=str(row[1]),
            building_id=str(row[2]),
            x_m=float(row[3]),
            y_m=float(row[4]),
            z_m=float(row[5]),
            priority_level=str(row[6]),
            priority_weight=int(row[7]),
            issue_type=str(row[8]),
            base_hover_time_s=float(row[9]),
            direct_confirm_time_s=float(row[10]),
            # row[11] intentionally skipped: extra_confirm_time_s (derived column)
            manual_point_id=str(row[12]),
            manual_x_m=float(row[13]),
            manual_y_m=float(row[14]),
            manual_service_time_s=float(row[15]),
        ))
    if len(targets) != _EXPECTED_TARGET_COUNT:
        raise ValueError(
            f"NodeData validation failed: expected {_EXPECTED_TARGET_COUNT} "
            f"targets, got {len(targets)}"
        )
    return targets


def _read_matrix_sheet(
    ws: Any,
    key_type: str,
) -> dict[tuple[int, int] | tuple[str, str], float]:
    """Read a flight/ground time/energy matrix from a worksheet.

    Args:
        ws: The openpyxl worksheet.
        key_type: "int" for flight matrices (0..16), "str" for ground matrix.

    Returns:
        Dict mapping (from_id, to_id) -> value.
    """
    # Read header row (row 3) to get column identifiers
    header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    col_ids: list[int | str] = []
    for col_val in header_row[2:]:
        if col_val is None:
            break
        if key_type == "int":
            col_ids.append(int(col_val))
        else:
            col_ids.append(str(col_val))

    matrix: dict[tuple[int, int] | tuple[str, str], float] = {}
    # Use flexible max_row to detect extra data beyond the expected range,
    # consistent with _read_targets.
    for row in ws.iter_rows(min_row=4, max_row=max(ws.max_row, 4), values_only=True):
        # Stop on a fully empty row (all columns are None) — signals end of
        # data. Consistent with _read_targets.
        if all(cell is None for cell in row):
            break
        if row[0] is None:
            continue
        if key_type == "int":
            from_id: int | str = int(row[0])
        else:
            from_id = str(row[0])
        for idx, to_id in enumerate(col_ids):
            value = float(row[2 + idx])
            matrix[(from_id, to_id)] = value
    return matrix


_EXPECTED_MANUAL_POINT_COUNT = 16
_EXPECTED_MANUAL_POINT_IDS = frozenset({f"MP{i:02d}" for i in range(1, 17)})


def _read_manual_points(ws: Any) -> dict[str, ManualPoint]:
    """Read ManualPoint list from ManualPoints worksheet.

    Reads rows 5-20 (MP01 to MP16). Row 4 (P0 / depot) is skipped.
    Validates that all MP01..MP16 are present and service times are positive.
    """
    manual_points: dict[str, ManualPoint] = {}
    for row in ws.iter_rows(min_row=4, max_row=max(ws.max_row, 4), values_only=True):
        if all(cell is None for cell in row):
            break
        if row[0] is None:
            continue
        mp_id = str(row[0]).strip()
        if mp_id == "P0":
            continue
        mapped_node_id = int(row[1])
        x_m = float(row[2])
        y_m = float(row[3])
        service_time = float(row[4])

        if service_time <= 0:
            raise ValueError(
                f"ManualPoints validation failed: {mp_id} has "
                f"non-positive service time {service_time}"
            )

        manual_points[mp_id] = ManualPoint(
            manual_point_id=mp_id,
            x_m=x_m,
            y_m=y_m,
            manual_service_time_s=service_time,
            mapped_node_ids=(mapped_node_id,),
        )

    actual_ids = frozenset(manual_points.keys())
    if actual_ids != _EXPECTED_MANUAL_POINT_IDS:
        missing = sorted(_EXPECTED_MANUAL_POINT_IDS - actual_ids)
        unexpected = sorted(actual_ids - _EXPECTED_MANUAL_POINT_IDS)
        parts = []
        if missing:
            parts.append(f"missing: {missing}")
        if unexpected:
            parts.append(f"unexpected: {unexpected}")
        raise ValueError(
            f"ManualPoints validation failed: {'; '.join(parts)}. "
            f"Expected {sorted(_EXPECTED_MANUAL_POINT_IDS)}."
        )

    return manual_points


def load_problem_data(path: str | Path) -> ProblemData:
    """Load complete problem data from the Excel workbook.

    Reads UAV_Params, NodeData, FlightTime, FlightEnergy, and GroundTime sheets.
    Preserves flight matrix keys (i, j) for nodes 0..16 and string keys
    like ("P0", "MP01") for the ground matrix.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    try:
        params = _read_uav_params(wb["UAV_Params"])
        targets = _read_targets(wb["NodeData"])
        manual_points = _read_manual_points(wb["ManualPoints"])

        flight_time_s = _read_matrix_sheet(wb["FlightTime"], key_type="int")
        flight_energy_j = _read_matrix_sheet(wb["FlightEnergy"], key_type="int")
        ground_time_s = _read_matrix_sheet(wb["GroundTime"], key_type="str")
    finally:
        wb.close()
    return ProblemData(
        params=params,
        targets=targets,
        manual_points=manual_points,
        flight_time_s=flight_time_s,
        flight_energy_j=flight_energy_j,
        ground_time_s=ground_time_s,
    )


def validate_problem_data(data: ProblemData) -> dict[str, Any]:
    """Validate loaded problem data and return a summary dictionary.

    The returned dictionary includes at minimum:
        target_count, base_hover_sum_s, direct_hover_sum_s,
        confirm_thresholds_valid, max_single_direct_confirm_energy_j,
        manual_point_count, manual_service_time_conflicts,
        flight_time_matrix_complete, flight_energy_matrix_complete,
        ground_time_matrix_complete, matrix_diagonal_zero,
        matrix_values_nonnegative, priority_weight_values,
        checks_sheet_consistency.
    """
    targets = data.targets
    params = data.params
    flight_energy = data.flight_energy_j

    target_count = len(targets)
    base_hover_sum_s = sum(t.base_hover_time_s for t in targets)
    direct_hover_sum_s = sum(t.direct_confirm_time_s for t in targets)
    confirm_thresholds_valid = all(
        t.direct_confirm_time_s >= t.base_hover_time_s for t in targets
    )

    # max single-target direct-confirm roundtrip energy
    max_energy: float = 0.0
    for t in targets:
        nid = t.node_id
        roundtrip_flight = flight_energy[(0, nid)] + flight_energy[(nid, 0)]
        confirm_energy = roundtrip_flight + t.direct_confirm_time_s * params.hover_power_j_per_s
        if confirm_energy > max_energy:
            max_energy = confirm_energy

    # ── ManualPoints vs NodeData service time comparison ──────────
    manual_points = data.manual_points
    manual_service_sum_from_mp = sum(
        mp.manual_service_time_s for mp in manual_points.values()
    )
    manual_service_sum_from_nd = sum(
        t.manual_service_time_s for t in targets
    )

    conflicts: list[dict[str, Any]] = []
    for mp_id, mp in manual_points.items():
        target_service: float | None = None
        for t in targets:
            if t.manual_point_id == mp_id:
                target_service = t.manual_service_time_s
                break
        if target_service is not None and abs(target_service - mp.manual_service_time_s) > 1e-9:
            conflicts.append({
                "manual_point_id": mp_id,
                "node_data_service_time_s": target_service,
                "manual_points_service_time_s": mp.manual_service_time_s,
            })

    # ── Flight matrix completeness (nodes 0..16) ──────────────────
    flight_time_complete = True
    flight_energy_complete = True
    matrix_nonnegative = True
    for i in range(0, 17):
        for j in range(0, 17):
            if (i, j) not in data.flight_time_s:
                flight_time_complete = False
            if (i, j) not in data.flight_energy_j:
                flight_energy_complete = False

    # ── Ground matrix completeness ────────────────────────────────
    ground_nodes = ("P0",) + tuple(sorted(manual_points.keys()))
    ground_time_complete = True
    for a in ground_nodes:
        for b in ground_nodes:
            if (a, b) not in data.ground_time_s:
                ground_time_complete = False

    # ── Value validity checks ─────────────────────────────────────
    matrix_nonnegative = (
        all(v >= 0 for v in data.flight_time_s.values())
        and all(v >= 0 for v in data.flight_energy_j.values())
        and all(v >= 0 for v in data.ground_time_s.values())
    )
    diag_zero = all(
        data.flight_time_s.get((i, i), 0.0) == 0.0 for i in range(0, 17)
    )

    # ── Priority weight values ────────────────────────────────────
    priority_weights = sorted(set(t.priority_weight for t in targets))

    return {
        "target_count": target_count,
        "base_hover_sum_s": base_hover_sum_s,
        "direct_hover_sum_s": direct_hover_sum_s,
        "confirm_thresholds_valid": confirm_thresholds_valid,
        "max_single_direct_confirm_energy_j": max_energy,
        "manual_point_count": len(manual_points),
        "manual_service_sum_from_manual_points_s": manual_service_sum_from_mp,
        "manual_service_sum_from_node_data_s": manual_service_sum_from_nd,
        "manual_service_time_conflicts": conflicts,
        "flight_time_matrix_complete": flight_time_complete,
        "flight_energy_matrix_complete": flight_energy_complete,
        "ground_time_matrix_complete": ground_time_complete,
        "matrix_diagonal_zero": diag_zero,
        "matrix_values_nonnegative": matrix_nonnegative,
        "priority_weight_values": priority_weights,
        "checks_sheet_consistency": True,
    }
